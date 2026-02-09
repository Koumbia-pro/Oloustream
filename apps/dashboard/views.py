from datetime import timedelta

from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.forms import AdminPasswordChangeForm
from django.db.models import Sum
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone

from apps.accounts.models import User
from apps.accounts.forms import EmployeeCreateForm, EmployeeUpdateForm
from apps.messaging.models import Message
from apps.notifications.models import Notification
from apps.payments.models import Payment, PaymentStatus
from apps.services_app.models import Service, Training, Partner
from apps.studio.choices import EquipmentStatus, ReservationStatus
from apps.studio.models import Reservation, Equipment, Studio
from apps.studio.forms import EquipmentForm, ReservationAdminForm
from apps.services_app.models import Service, Training, Partner, Offer
from apps.services_app.forms import ServiceForm, OfferForm, TrainingForm, PartnerForm
from apps.accounts.models import User, EmployeeProfile
from django.core.paginator import Paginator
from django.db.models import Q, Sum
from apps.accounts.models import User, EmployeeProfile, ContractTypeChoices
from django.http import HttpResponse
from openpyxl import Workbook
from apps.studio.models import Equipment, EquipmentCategory
from apps.studio.choices import EquipmentStatus
from openpyxl.styles import Font, Alignment, PatternFill
from apps.studio.models import Reservation, Equipment, Studio, ReservationStatusHistory
from apps.studio.services import log_reservation_status_change
from apps.studio.choices import ReservationStatus
from apps.services_app.models import Service, ServiceCategory, ServiceTypeChoices
from apps.services_app.models import Offer
from apps.services_app.models import Training, TrainingCategory, TrainingLevelChoices, TrainingModeChoices
from apps.services_app.forms import TrainingForm
from apps.notifications.services import notify_user_reservation_status_change
from apps.notifications.models import Notification, NotificationTypeChoices
from apps.studio.forms import StudioForm
from django.views.decorators.http import require_POST
from django.shortcuts import redirect
from django.contrib import messages
from apps.notifications.emailing import send_reservation_status_changed_email



# ---------- DASHBOARD GLOBAL ----------
# apps/dashboard/views.py

from datetime import timedelta
from django.db.models import Sum, Count, Q
from calendar import monthrange

from apps.accounts.models import User
from apps.business_partners.models import (
    BusinessPartner, PartnerApplication, Contract, CommissionPayment
)
from apps.messaging.models import Message, Conversation
from apps.notifications.models import Notification, NotificationTypeChoices
from apps.payments.models import Payment, PaymentStatus
from apps.services_app.models import Service, Training, Partner, JobOffer, JobApplication
from apps.studio.models import Reservation, Equipment, Studio
from apps.studio.choices import EquipmentStatus, ReservationStatus


@staff_member_required
def dashboard_view(request):
    """
    Dashboard principal avec toutes les statistiques et graphiques.
    """
    
    # ==================== PÉRIODE DE FILTRAGE ====================
    period = request.GET.get('period', 'all')  # all, today, week, month, year
    
    now = timezone.now()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    
    if period == 'today':
        start_date = today_start
        end_date = now
    elif period == 'week':
        start_date = today_start - timedelta(days=today_start.weekday())
        end_date = now
    elif period == 'month':
        start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        end_date = now
    elif period == 'year':
        start_date = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        end_date = now
    else:
        start_date = None
        end_date = None
    
    # ==================== STATS PRINCIPALES ====================
    
    # Chiffre d'affaires
    total_revenue_qs = Payment.objects.filter(status=PaymentStatus.PAID)
    if start_date:
        total_revenue_qs = total_revenue_qs.filter(created_at__gte=start_date, created_at__lte=end_date)
    total_revenue = total_revenue_qs.aggregate(total=Sum('amount'))['total'] or 0
    
    # Réservations
    reservations_qs = Reservation.objects.all()
    if start_date:
        reservations_qs = reservations_qs.filter(created_at__gte=start_date, created_at__lte=end_date)
    total_reservations = reservations_qs.count()
    
    # Clients
    total_clients = User.objects.filter(role=User.Role.CLIENT).count()
    
    # Employés
    total_employees = User.objects.filter(is_employee=True, is_active=True).count()
    
    # Équipements
    equipments_available = Equipment.objects.filter(status=EquipmentStatus.AVAILABLE).count()
    equipments_maintenance = Equipment.objects.filter(status=EquipmentStatus.MAINTENANCE).count()
    total_equipments = Equipment.objects.count()
    
    # Services & Formations
    active_services = Service.objects.filter(is_active=True).count()
    active_trainings = Training.objects.filter(is_active=True).count()
    
    # Partenaires simples
    active_partners = Partner.objects.filter(active=True).count()
    
    # Revenus récents (30 derniers jours)
    thirty_days_ago = now - timedelta(days=30)
    recent_revenue = Payment.objects.filter(
        status=PaymentStatus.PAID,
        created_at__gte=thirty_days_ago
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    # ==================== BUSINESS PARTNERS ====================
    total_business_partners = BusinessPartner.objects.filter(is_active=True).count()
    pending_applications = PartnerApplication.objects.filter(status='pending').count()
    pending_contracts = Contract.objects.filter(status='pending').count()
    
    bp_stats = BusinessPartner.objects.aggregate(
        total_commission_earned=Sum('total_commission_earned'),
        total_commission_paid=Sum('total_commission_paid'),
    )
    total_commission_earned = bp_stats['total_commission_earned'] or 0
    total_commission_paid = bp_stats['total_commission_paid'] or 0
    pending_commission = total_commission_earned - total_commission_paid
    
    # ==================== RÉSERVATIONS PAR STATUT ====================
    reservations_by_status = {
        'pending': Reservation.objects.filter(status=ReservationStatus.PENDING).count(),
        'confirmed': Reservation.objects.filter(status=ReservationStatus.CONFIRMED).count(),
        'completed': Reservation.objects.filter(status=ReservationStatus.COMPLETED).count(),
        'cancelled': Reservation.objects.filter(status=ReservationStatus.CANCELLED).count(),
        'rejected': Reservation.objects.filter(status=ReservationStatus.REJECTED).count(),
    }
    
    # ==================== TOP 5 STUDIOS ====================
    top_studios = (
        Reservation.objects
        .filter(studio__isnull=False)
        .values('studio__name')
        .annotate(count=Count('id'))
        .order_by('-count')[:5]
    )
    
    # ==================== TOP 5 SERVICES ====================
    top_services = (
        Reservation.objects
        .filter(service__isnull=False)
        .values('service__name')
        .annotate(count=Count('id'))
        .order_by('-count')[:5]
    )
    
    # ==================== TOP 5 PARTENAIRES D'AFFAIRES ====================
    top_partners = (
        BusinessPartner.objects
        .filter(is_active=True)
        .order_by('-total_revenue')[:5]
    )
    
    # ==================== GRAPHIQUE ÉVOLUTION CA (6 derniers mois) ====================
    revenue_chart_data = []
    for i in range(5, -1, -1):
        month_date = now - timedelta(days=30 * i)
        month_start = month_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        last_day = monthrange(month_start.year, month_start.month)[1]
        month_end = month_start.replace(day=last_day, hour=23, minute=59, second=59)
        
        month_revenue = Payment.objects.filter(
            status=PaymentStatus.PAID,
            created_at__gte=month_start,
            created_at__lte=month_end
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        revenue_chart_data.append({
            'month': month_start.strftime('%b %Y'),
            'revenue': float(month_revenue)
        })
    
    # ==================== ALERTES & ACTIONS URGENTES ====================
    alerts = []
    
    # Réservations en attente
    pending_reservations_count = reservations_by_status['pending']
    if pending_reservations_count > 0:
        alerts.append({
            'icon': 'ph-calendar-check',
            'color': 'warning',
            'text': f"{pending_reservations_count} réservation{'s' if pending_reservations_count > 1 else ''} en attente de confirmation",
            'link': '/dashboard/reservations/?status=PENDING',
        })
    
    # Candidatures partenaires
    if pending_applications > 0:
        alerts.append({
            'icon': 'ph-user-plus',
            'color': 'info',
            'text': f"{pending_applications} candidature{'s' if pending_applications > 1 else ''} partenaire{'s' if pending_applications > 1 else ''} à examiner",
            'link': '/dashboard/partenaires/candidatures/?status=pending',
        })
    
    # Contrats à valider
    if pending_contracts > 0:
        alerts.append({
            'icon': 'ph-file-text',
            'color': 'primary',
            'text': f"{pending_contracts} contrat{'s' if pending_contracts > 1 else ''} à valider",
            'link': '/dashboard/partenaires/contrats/?status=pending',
        })
    
    # Messages non lus
    unread_messages_count = Notification.objects.filter(
        user=request.user,
        notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
        is_read=False
    ).count()
    if unread_messages_count > 0:
        alerts.append({
            'icon': 'ph-chat-circle-dots',
            'color': 'success',
            'text': f"{unread_messages_count} message{'s' if unread_messages_count > 1 else ''} non lu{'s' if unread_messages_count > 1 else ''}",
            'link': '/messaging/admin/conversations/',
        })
    
    # Équipements en maintenance
    if equipments_maintenance > 0:
        alerts.append({
            'icon': 'ph-wrench',
            'color': 'danger',
            'text': f"{equipments_maintenance} équipement{'s' if equipments_maintenance > 1 else ''} en maintenance",
            'link': '/dashboard/equipments/?status=MAINTENANCE',
        })
    
    # Candidatures emploi/stage
    pending_job_applications = JobApplication.objects.filter(status='PENDING').count()
    if pending_job_applications > 0:
        alerts.append({
            'icon': 'ph-briefcase',
            'color': 'info',
            'text': f"{pending_job_applications} candidature{'s' if pending_job_applications > 1 else ''} emploi/stage à traiter",
            'link': '/services/admin/jobs/',
        })
    
    # ==================== DERNIÈRES RÉSERVATIONS ====================
    last_reservations = (
        Reservation.objects
        .select_related('user', 'studio', 'service')
        .order_by('-created_at')[:10]
    )
    
    # ==================== FORMATIONS À VENIR ====================
    upcoming_trainings = Training.objects.filter(
        is_active=True,
        start_date__gte=now.date()
    ).order_by('start_date')[:5]
    
    # ==================== OFFRES D'EMPLOI OUVERTES ====================
    open_job_offers = JobOffer.objects.filter(
        status='PUBLISHED',
        deadline__gte=now.date()
    ).count()
    
    # ==================== STUDIOS (OCCUPATION) ====================
    total_studios = Studio.objects.filter(is_active=True).count()
    
    # Calcul taux d'occupation (réservations confirmées sur période)
    if start_date and total_studios > 0:
        confirmed_reservations = Reservation.objects.filter(
            status=ReservationStatus.CONFIRMED,
            start_datetime__gte=start_date,
            start_datetime__lte=end_date
        ).count()
        # Simplification : on considère qu'un studio peut être réservé 8h/jour
        available_slots = total_studios * (end_date - start_date).days * 8
        occupation_rate = (confirmed_reservations / available_slots * 100) if available_slots > 0 else 0
    else:
        occupation_rate = 0
    
    # ==================== NOTIFICATIONS ====================
    unread_notifications_count = Notification.objects.filter(
        user=request.user,
        is_read=False
    ).count()
    
    # ==================== CONTEXT ====================
    context = {
        # Filtres
        'current_period': period,
        
        # Stats principales
        'total_revenue': total_revenue,
        'total_reservations': total_reservations,
        'total_clients': total_clients,
        'total_employees': total_employees,
        'equipments_available': equipments_available,
        'equipments_maintenance': equipments_maintenance,
        'total_equipments': total_equipments,
        'active_services': active_services,
        'active_trainings': active_trainings,
        'active_partners': active_partners,
        'recent_revenue': recent_revenue,
        
        # Business Partners
        'total_business_partners': total_business_partners,
        'pending_applications': pending_applications,
        'pending_contracts': pending_contracts,
        'total_commission_earned': total_commission_earned,
        'total_commission_paid': total_commission_paid,
        'pending_commission': pending_commission,
        
        # Réservations
        'reservations_by_status': reservations_by_status,
        
        # Top performers
        'top_studios': top_studios,
        'top_services': top_services,
        'top_partners': top_partners,
        
        # Graphiques
        'revenue_chart_data': revenue_chart_data,
        
        # Alertes
        'alerts': alerts,
        
        # Listes
        'last_reservations': last_reservations,
        'upcoming_trainings': upcoming_trainings,
        'open_job_offers': open_job_offers,
        
        # Studios
        'total_studios': total_studios,
        'occupation_rate': round(occupation_rate, 1),
        
        # Notifications
        'unread_notifications_count': unread_notifications_count,
        'unread_messages_count': unread_messages_count,
    }
    
    return render(request, "admin/dashboard.html", context)

# @staff_member_required
# def dashboard_view(request):
#     # Chiffres globaux
#     total_revenue = Payment.objects.filter(
#         status=PaymentStatus.PAID
#     ).aggregate(total=Sum('amount'))['total'] or 0

#     total_reservations = Reservation.objects.count()
#     total_clients = User.objects.filter(role=User.Role.CLIENT).count()
#     total_employees = User.objects.filter(is_employee=True).count()

#     from apps.studio.choices import EquipmentStatus

#     equipments_available = Equipment.objects.filter(status=EquipmentStatus.AVAILABLE).count()
#     equipments_maintenance = Equipment.objects.filter(status=EquipmentStatus.MAINTENANCE).count()

#     active_services = Service.objects.filter(is_active=True).count()
#     active_trainings = Training.objects.filter(is_active=True).count()
#     active_partners = Partner.objects.filter(active=True).count()

#     # Revenus récents (30 derniers jours)
#     thirty_days_ago = timezone.now() - timedelta(days=30)
#     recent_revenue = Payment.objects.filter(
#         status=PaymentStatus.PAID,
#         created_at__gte=thirty_days_ago
#     ).aggregate(total=Sum('amount'))['total'] or 0

#     # Dernières réservations
#     last_reservations = Reservation.objects.select_related('user', 'studio').order_by('-created_at')[:5]

#     # Notifications & messages
#     unread_notifications_count = Notification.objects.filter(
#         user=request.user,
#         is_read=False
#     ).count()

#     unread_messages_count = Notification.objects.filter(
#         user=request.user,
#         notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
#         is_read=False
#     ).count(),

#     context = {
#         "total_revenue": total_revenue,
#         "total_reservations": total_reservations,
#         "total_clients": total_clients,
#         "total_employees": total_employees,
#         "equipments_available": equipments_available,
#         "equipments_maintenance": equipments_maintenance,
#         "active_services": active_services,
#         "active_trainings": active_trainings,
#         "active_partners": active_partners,
#         "recent_revenue": recent_revenue,
#         "last_reservations": last_reservations,
#         "unread_notifications_count": unread_notifications_count,
#         "unread_messages_count": unread_messages_count,
#     }
#     return render(request, "admin/dashboard.html", context)


# ---------- GESTION DES EMPLOYÉS ----------

@staff_member_required
def employee_list_view(request):
    qs = (
        User.objects
        .filter(is_employee=True)
        .select_related('employee_profile')
        .order_by('id')
    )

    # --- Recherche libre ---
    q = request.GET.get('q', '').strip()

    if q:
        qs = qs.filter(
            Q(username__icontains=q) |
            Q(first_name__icontains=q) |
            Q(last_name__icontains=q) |
            Q(email__icontains=q) |
            Q(phone__icontains=q) |
            Q(employee_profile__position__icontains=q) |
            Q(employee_profile__city__icontains=q)
        )

    # --- Filtres ---
    selected_role = request.GET.get('role', '')
    selected_status = request.GET.get('status', '')          # active / inactive
    selected_contract_type = request.GET.get('contract_type', '')

    if selected_role:
        qs = qs.filter(role=selected_role)

    if selected_status == 'active':
        qs = qs.filter(is_active=True)
    elif selected_status == 'inactive':
        qs = qs.filter(is_active=False)

    if selected_contract_type:
        qs = qs.filter(employee_profile__contract_type=selected_contract_type)

    # --- Pagination ---
    paginator = Paginator(qs, 10)  # 10 employés par page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Conserver les filtres dans la pagination
    filters_qd = request.GET.copy()
    if 'page' in filters_qd:
        filters_qd.pop('page')
    base_querystring = filters_qd.urlencode()

    context = {
        "page_obj": page_obj,
        "q": q,
        "selected_role": selected_role,
        "selected_status": selected_status,
        "selected_contract_type": selected_contract_type,
        "role_choices": User.Role.choices,
        "contract_type_choices": ContractTypeChoices.choices,
        "base_querystring": base_querystring,
        "unread_notifications_count": Notification.objects.filter(user=request.user, is_read=False).count(),
        "unread_messages_count": Notification.objects.filter(
            user=request.user,
            notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
            is_read=False
        ).count(),
    }
    return render(request, "admin/employees/list.html", context)


@staff_member_required
def employee_export_excel_view(request):
    """
    Exporte la liste des employés filtrée au format Excel (XLSX).
    """
    qs = (
        User.objects
        .filter(is_employee=True)
        .select_related('employee_profile')
        .order_by('id')
    )

    # --- mêmes filtres que la liste ---
    q = request.GET.get('q', '').strip()
    selected_role = request.GET.get('role', '')
    selected_status = request.GET.get('status', '')
    selected_contract_type = request.GET.get('contract_type', '')

    if q:
        qs = qs.filter(
            Q(username__icontains=q) |
            Q(first_name__icontains=q) |
            Q(last_name__icontains=q) |
            Q(email__icontains=q) |
            Q(phone__icontains=q) |
            Q(employee_profile__position__icontains=q) |
            Q(employee_profile__city__icontains=q)
        )

    if selected_role:
        qs = qs.filter(role=selected_role)

    if selected_status == 'active':
        qs = qs.filter(is_active=True)
    elif selected_status == 'inactive':
        qs = qs.filter(is_active=False)

    if selected_contract_type:
        qs = qs.filter(employee_profile__contract_type=selected_contract_type)

    # --- Création du fichier Excel ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Employés"

    # En-têtes
    headers = [
        "ID",
        "Nom d'utilisateur",
        "Nom",
        "Prénom",
        "Email",
        "Téléphone",
        "Rôle",
        "Date d'embauche",
        "Ville",
        "Poste",
        "Type de contrat",
        "Salaire",
    ]
    ws.append(headers)

    # Lignes de données
    for e in qs:
        profile = getattr(e, 'employee_profile', None)

        hire_date = profile.hire_date.strftime("%d/%m/%Y") if profile and profile.hire_date else ""
        city = profile.city if profile else ""
        position = profile.position if profile else ""
        contract_type = profile.get_contract_type_display() if profile and profile.contract_type else ""
        salary = float(profile.salary) if profile and profile.salary is not None else None

        ws.append([
            e.id,
            e.username,
            e.last_name,
            e.first_name,
            e.email,
            e.phone,
            e.get_role_display(),
            hire_date,
            city,
            position,
            contract_type,
            salary,
        ])

    # Style minimal des en-têtes
    from openpyxl.styles import Font, Alignment, PatternFill

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="111827")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Ajuster un peu la largeur de colonnes
    column_widths = [6, 18, 18, 18, 26, 16, 14, 14, 16, 20, 18, 12]
    for i, column_cells in enumerate(ws.columns, start=1):
        try:
            ws.column_dimensions[column_cells[0].column_letter].width = column_widths[i-1]
        except IndexError:
            pass

    # Réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    timestamp = timezone.now().strftime("%Y%m%d_%H%M")
    filename = f"employes_oloustream_{timestamp}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response

@staff_member_required
def employee_detail_view(request, user_id):
    employee = get_object_or_404(
        User.objects.select_related('employee_profile'),
        pk=user_id,
        is_employee=True
    )

    profile = getattr(employee, 'employee_profile', None)

    context = {
        "employee": employee,
        "profile": profile,
        "unread_notifications_count": Notification.objects.filter(user=request.user, is_read=False).count(),
        "unread_messages_count": Notification.objects.filter(
            user=request.user,
            notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
            is_read=False
        ).count(),
    }
    return render(request, "admin/employees/detail.html", context)

@staff_member_required
def employee_create_view(request):
    if request.method == "POST":
        form = EmployeeCreateForm(request.POST, request.FILES)  # <-- + request.FILES
        if form.is_valid():
            user = form.save()
            messages.success(request, f"Employé {user.username} créé avec succès.")
            return redirect('dashboard:employees_list')
    else:
        form = EmployeeCreateForm()

    context = {
        "form": form,
        "unread_notifications_count": Notification.objects.filter(user=request.user, is_read=False).count(),
        "unread_messages_count": Notification.objects.filter(
            user=request.user,
            notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
            is_read=False
        ).count(),
    }
    return render(request, "admin/employees/form.html", context)


@staff_member_required
def employee_update_view(request, user_id):
    employee = get_object_or_404(User, pk=user_id, is_employee=True)

    if request.method == "POST":
        form = EmployeeUpdateForm(request.POST, request.FILES, instance=employee)  # <-- + request.FILES
        if form.is_valid():
            form.save()
            messages.success(request, "Employé mis à jour avec succès.")
            return redirect('dashboard:employees_list')
    else:
        form = EmployeeUpdateForm(instance=employee)

    context = {
        "form": form,
        "employee": employee,
        "unread_notifications_count": Notification.objects.filter(user=request.user, is_read=False).count(),
        "unread_messages_count": Notification.objects.filter(
            user=request.user,
            notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
            is_read=False
        ).count(),
    }
    return render(request, "admin/employees/form.html", context)


@staff_member_required
def employee_delete_view(request, user_id):
    employee = get_object_or_404(User, pk=user_id, is_employee=True)

    if request.method == "POST":
        username = employee.username
        employee.delete()
        messages.success(request, f"Employé {username} supprimé.")
        return redirect('dashboard:employees_list')

    context = {
        "employee": employee,
        "unread_notifications_count": Notification.objects.filter(user=request.user, is_read=False).count(),
        "unread_messages_count": Notification.objects.filter(
            user=request.user,
            notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
            is_read=False
        ).count(),
    }
    return render(request, "admin/employees/confirm_delete.html", context)


@staff_member_required
def employee_change_password_view(request, user_id):
    employee = get_object_or_404(User, pk=user_id, is_employee=True)

    if request.method == "POST":
        form = AdminPasswordChangeForm(employee, request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Mot de passe modifié avec succès.")
            return redirect('dashboard:employees_list')
    else:
        form = AdminPasswordChangeForm(employee)

    context = {
        "form": form,
        "employee": employee,
       "unread_notifications_count": Notification.objects.filter(user=request.user, is_read=False).count(),
        "unread_messages_count": Notification.objects.filter(
            user=request.user,
            notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
            is_read=False
        ).count(),
    }
    return render(request, "admin/employees/change_password.html", context)


# ---------- GESTION DES ÉQUIPEMENTS ----------

@staff_member_required
def equipment_list_view(request):
    qs = (
        Equipment.objects
        .select_related('category', 'current_user')
        .order_by('id')
    )

    # --- Recherche libre ---
    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(
            Q(name__icontains=q) |
            Q(brand__icontains=q) |
            Q(model__icontains=q) |
            Q(serial_number__icontains=q) |
            Q(location__icontains=q) |
            Q(category__name__icontains=q)
        )

    # --- Filtres ---
    selected_category = request.GET.get('category', '')
    selected_status = request.GET.get('status', '')
    selected_rental = request.GET.get('rental', '')  # yes / no

    if selected_category:
        qs = qs.filter(category_id=selected_category)

    if selected_status:
        qs = qs.filter(status=selected_status)

    if selected_rental == 'yes':
        qs = qs.filter(is_available_for_rent=True)
    elif selected_rental == 'no':
        qs = qs.filter(is_available_for_rent=False)

    # --- Statistiques sur la liste filtrée ---
    stats_qs = qs  # la même requête filtrée

    total_count = stats_qs.count()
    available_count = stats_qs.filter(status=EquipmentStatus.AVAILABLE).count()
    in_use_count = stats_qs.filter(status=EquipmentStatus.IN_USE).count()
    maintenance_count = stats_qs.filter(status=EquipmentStatus.MAINTENANCE).count()
    out_of_service_count = stats_qs.filter(status=EquipmentStatus.OUT_OF_SERVICE).count()
    retired_count = stats_qs.filter(status=EquipmentStatus.RETIRED).count()
    total_value = stats_qs.aggregate(Sum('purchase_price'))['purchase_price__sum'] or 0

    # --- Pagination ---
    paginator = Paginator(qs, 10)  # 10 équipements par page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Conserver les filtres dans la pagination et l’export
    filters_qd = request.GET.copy()
    if 'page' in filters_qd:
        filters_qd.pop('page')
    base_querystring = filters_qd.urlencode()

    categories = EquipmentCategory.objects.all().order_by('name')

    context = {
        "page_obj": page_obj,
        "q": q,
        "selected_category": selected_category,
        "selected_status": selected_status,
        "selected_rental": selected_rental,
        "categories": categories,
        "status_choices": EquipmentStatus.choices,
        "base_querystring": base_querystring,

        # Stats
        "total_count": total_count,
        "available_count": available_count,
        "in_use_count": in_use_count,
        "maintenance_count": maintenance_count,
        "out_of_service_count": out_of_service_count,
        "retired_count": retired_count,
        "total_value": total_value,

       "unread_notifications_count": Notification.objects.filter(user=request.user, is_read=False).count(),
        "unread_messages_count": Notification.objects.filter(
            user=request.user,
            notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
            is_read=False
        ).count(),
    }
    return render(request, "admin/equipments/list.html", context)


@staff_member_required
def equipment_export_excel_view(request):
    """
    Exporte la liste des équipements filtrée au format Excel (XLSX).
    """
    qs = (
        Equipment.objects
        .select_related('category', 'current_user')
        .order_by('id')
    )

    # mêmes filtres que la liste
    q = request.GET.get('q', '').strip()
    selected_category = request.GET.get('category', '')
    selected_status = request.GET.get('status', '')
    selected_rental = request.GET.get('rental', '')

    if q:
        qs = qs.filter(
            Q(name__icontains=q) |
            Q(brand__icontains=q) |
            Q(model__icontains=q) |
            Q(serial_number__icontains=q) |
            Q(location__icontains=q) |
            Q(category__name__icontains=q)
        )

    if selected_category:
        qs = qs.filter(category_id=selected_category)

    if selected_status:
        qs = qs.filter(status=selected_status)

    if selected_rental == 'yes':
        qs = qs.filter(is_available_for_rent=True)
    elif selected_rental == 'no':
        qs = qs.filter(is_available_for_rent=False)

    # Création du classeur Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Équipements"

    headers = [
        "ID",
        "Nom",
        "Catégorie",
        "Marque",
        "Modèle",
        "N° de série",
        "Statut",
        "Disponible à la location",
        "Emplacement",
        "Utilisateur actuel",
        "Date d'achat",
        "Prix d'achat",
        "Âge (années)",
        "Dernière maintenance",
        "Prochaine maintenance",
    ]
    ws.append(headers)

    for e in qs:
        profile_age = e.age_years if e.age_years is not None else ""
        current_user_name = e.current_user.get_full_name() if e.current_user else ""
        purchase_date = e.purchase_date.strftime("%d/%m/%Y") if e.purchase_date else ""
        purchase_price = float(e.purchase_price) if e.purchase_price is not None else None
        last_maintenance = e.last_maintenance_date.strftime("%d/%m/%Y") if e.last_maintenance_date else ""
        next_maintenance = e.next_maintenance_date.strftime("%d/%m/%Y") if e.next_maintenance_date else ""

        ws.append([
            e.id,
            e.name,
            e.category.name if e.category else "",
            e.brand,
            e.model,
            e.serial_number,
            e.get_status_display(),
            "Oui" if e.is_available_for_rent else "Non",
            e.location,
            current_user_name,
            purchase_date,
            purchase_price,
            profile_age,
            last_maintenance,
            next_maintenance,
        ])

    # Style des en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="111827")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Largeurs de colonnes
    column_widths = [6, 24, 18, 16, 16, 18, 16, 10, 18, 22, 14, 14, 10, 18, 18]
    for i, column_cells in enumerate(ws.columns, start=1):
        try:
            ws.column_dimensions[column_cells[0].column_letter].width = column_widths[i-1]
        except IndexError:
            pass

    # Réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    timestamp = timezone.now().strftime("%Y%m%d_%H%M")
    filename = f"equipements_oloustream_{timestamp}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


@staff_member_required
def equipment_detail_view(request, equipment_id):
    equipment = get_object_or_404(
        Equipment.objects.select_related('category', 'current_user'),
        pk=equipment_id
    )

    context = {
        "equipment": equipment,
        "unread_notifications_count": Notification.objects.filter(user=request.user, is_read=False).count(),
        "unread_messages_count": Notification.objects.filter(
            user=request.user,
            notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
            is_read=False
        ).count(),
    }
    return render(request, "admin/equipments/detail.html", context)


@staff_member_required
def equipment_create_view(request):
    if request.method == "POST":
        form = EquipmentForm(request.POST, request.FILES)  # <-- + request.FILES
        if form.is_valid():
            equipment = form.save()
            messages.success(request, f"Équipement {equipment.name} créé avec succès.")
            return redirect('dashboard:equipments_list')
    else:
        form = EquipmentForm()

    context = {
        "form": form,
        "equipment": None,
        "unread_notifications_count": Notification.objects.filter(user=request.user, is_read=False).count(),
        "unread_messages_count": Notification.objects.filter(
            user=request.user,
            notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
            is_read=False
        ).count(),
    }
    return render(request, "admin/equipments/form.html", context)


@staff_member_required
def equipment_update_view(request, equipment_id):
    equipment = get_object_or_404(Equipment, pk=equipment_id)

    if request.method == "POST":
        form = EquipmentForm(request.POST, request.FILES, instance=equipment)  # <-- + request.FILES
        if form.is_valid():
            form.save()
            messages.success(request, "Équipement mis à jour avec succès.")
            return redirect('dashboard:equipments_list')
    else:
        form = EquipmentForm(instance=equipment)

    context = {
        "form": form,
        "equipment": equipment,
        "unread_notifications_count": Notification.objects.filter(user=request.user, is_read=False).count(),
        "unread_messages_count": Notification.objects.filter(
            user=request.user,
            notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
            is_read=False
        ).count(),
    }
    return render(request, "admin/equipments/form.html", context)

@staff_member_required
def equipment_delete_view(request, equipment_id):
    equipment = get_object_or_404(Equipment, pk=equipment_id)

    if request.method == "POST":
        name = equipment.name
        equipment.delete()
        messages.success(request, f"Équipement {name} supprimé.")
        return redirect('dashboard:equipments_list')

    context = {
        "equipment": equipment,
        "unread_notifications_count": Notification.objects.filter(user=request.user, is_read=False).count(),
        "unread_messages_count": Notification.objects.filter(
            user=request.user,
            notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
            is_read=False
        ).count(),
    }
    return render(request, "admin/equipments/confirm_delete.html", context)


# ---------- GESTION DES RÉSERVATIONS ----------

@staff_member_required
def reservation_list_view(request):
    qs = (
        Reservation.objects
        .select_related('user', 'studio', 'service', 'assigned_technician')
        .prefetch_related('equipments')
        .order_by('-created_at')
    )

    # --- Recherche libre ---
    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(
            Q(user__username__icontains=q) |
            Q(user__first_name__icontains=q) |
            Q(user__last_name__icontains=q) |
            Q(user__email__icontains=q) |
            Q(studio__name__icontains=q) |
            Q(service__name__icontains=q) |
            Q(admin_comment__icontains=q)
        )

    # --- Filtres ---
    selected_status = request.GET.get('status', '')
    selected_studio = request.GET.get('studio', '')
    selected_service = request.GET.get('service', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    if selected_status:
        qs = qs.filter(status=selected_status)

    if selected_studio:
        qs = qs.filter(studio_id=selected_studio)

    if selected_service:
        qs = qs.filter(service_id=selected_service)

    if date_from:
        qs = qs.filter(start_datetime__date__gte=date_from)
    if date_to:
        qs = qs.filter(start_datetime__date__lte=date_to)

    # --- Statistiques sur la liste filtrée ---
    stats_qs = qs

    total_count = stats_qs.count()
    pending_count = stats_qs.filter(status=ReservationStatus.PENDING).count()
    confirmed_count = stats_qs.filter(status=ReservationStatus.CONFIRMED).count()
    completed_count = stats_qs.filter(status=ReservationStatus.COMPLETED).count()
    cancelled_count = stats_qs.filter(status=ReservationStatus.CANCELLED).count()
    rejected_count = stats_qs.filter(status=ReservationStatus.REJECTED).count()

    upcoming_count = stats_qs.filter(start_datetime__gte=timezone.now()).count()
    past_count = stats_qs.filter(end_datetime__lt=timezone.now()).count()

    # --- Pagination ---
    paginator = Paginator(qs, 10)  # 10 réservations par page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Conserver les filtres dans la pagination & export
    filters_qd = request.GET.copy()
    if 'page' in filters_qd:
        filters_qd.pop('page')
    base_querystring = filters_qd.urlencode()

    studios = Studio.objects.all().order_by('name')
    services = Service.objects.filter(is_active=True).order_by('name')

    context = {
        "page_obj": page_obj,
        "q": q,
        "selected_status": selected_status,
        "selected_studio": selected_studio,
        "selected_service": selected_service,
        "date_from": date_from,
        "date_to": date_to,
        "status_choices": ReservationStatus.choices,
        "studios": studios,
        "services": services,
        "base_querystring": base_querystring,

        # Stats
        "total_count": total_count,
        "pending_count": pending_count,
        "confirmed_count": confirmed_count,
        "completed_count": completed_count,
        "cancelled_count": cancelled_count,
        "rejected_count": rejected_count,
        "upcoming_count": upcoming_count,
        "past_count": past_count,

        "unread_notifications_count": Notification.objects.filter(user=request.user, is_read=False).count(),
        "unread_messages_count": Notification.objects.filter(
            user=request.user,
            notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
            is_read=False
        ).count(),
    }
    return render(request, "admin/reservations/list.html", context)

@staff_member_required
def reservation_export_excel_view(request):
    """
    Exporte la liste des réservations filtrée au format Excel (XLSX).
    """
    qs = (
        Reservation.objects
        .select_related('user', 'studio', 'service', 'assigned_technician')
        .prefetch_related('equipments')
        .order_by('-created_at')
    )

    # mêmes filtres que la liste
    q = request.GET.get('q', '').strip()
    selected_status = request.GET.get('status', '')
    selected_studio = request.GET.get('studio', '')
    selected_service = request.GET.get('service', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    if q:
        qs = qs.filter(
            Q(user__username__icontains=q) |
            Q(user__first_name__icontains=q) |
            Q(user__last_name__icontains=q) |
            Q(user__email__icontains=q) |
            Q(studio__name__icontains=q) |
            Q(service__name__icontains=q) |
            Q(admin_comment__icontains=q)
        )

    if selected_status:
        qs = qs.filter(status=selected_status)

    if selected_studio:
        qs = qs.filter(studio_id=selected_studio)

    if selected_service:
        qs = qs.filter(service_id=selected_service)

    if date_from:
        qs = qs.filter(start_datetime__date__gte=date_from)
    if date_to:
        qs = qs.filter(start_datetime__date__lte=date_to)

    # Génération Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Réservations"

    headers = [
        "ID",
        "Client",
        "Email client",
        "Téléphone client",
        "Service",
        "Studio",
        "Début",
        "Fin",
        "Statut",
        "Technicien assigné",
        "Nb équipements",
        "Créée le",
        "Commentaire admin",
    ]
    ws.append(headers)

    for r in qs:
        user = r.user
        technician = r.assigned_technician
        equipments_count = r.equipments.count()

        ws.append([
            r.id,
            user.get_full_name() or user.username,
            user.email,
            user.phone,
            r.service.name if r.service else "",
            r.studio.name if r.studio else "",
            r.start_datetime.strftime("%d/%m/%Y %H:%M") if r.start_datetime else "",
            r.end_datetime.strftime("%d/%m/%Y %H:%M") if r.end_datetime else "",
            r.get_status_display(),
            technician.get_full_name() if technician else "",
            equipments_count,
            r.created_at.strftime("%d/%m/%Y %H:%M") if r.created_at else "",
            (r.admin_comment or "")[:200],
        ])

    # Style en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="111827")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Largeurs colonnes
    column_widths = [6, 22, 24, 16, 22, 22, 18, 18, 14, 22, 14, 18, 40]
    for i, column_cells in enumerate(ws.columns, start=1):
        try:
            ws.column_dimensions[column_cells[0].column_letter].width = column_widths[i-1]
        except IndexError:
            pass

    # Réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    timestamp = timezone.now().strftime("%Y%m%d_%H%M")
    filename = f"reservations_oloustream_{timestamp}.xlsx"
    response['Content-Disposition'] = f'attachment; filename=\"{filename}\"'

    wb.save(response)
    return response


@staff_member_required
def reservation_quick_cancel_view(request, reservation_id):
    """
    Annulation rapide d'une réservation depuis la liste,
    avec enregistrement dans l'historique des statuts.
    """
    reservation = get_object_or_404(Reservation, pk=reservation_id)
    old_status = reservation.status
    new_status = ReservationStatus.CANCELLED

    if old_status != new_status:
        reservation.status = new_status
        reservation.save(update_fields=['status'])

        # Enregistrer l'historique avec une note explicative
        log_reservation_status_change(
            reservation=reservation,
            old_status=old_status,
            new_status=new_status,
            changed_by=request.user,
            note="Annulation rapide depuis la liste des réservations."
        )

        messages.success(request, f"La réservation #{reservation.id} a été annulée.")
    else:
        messages.info(request, "La réservation est déjà au statut annulé.")

    return redirect('dashboard:reservations_list')


@staff_member_required
def reservation_detail_view(request, reservation_id):
    reservation = get_object_or_404(
        Reservation.objects
        .select_related('user', 'studio', 'service', 'assigned_technician')
        .prefetch_related('equipments', 'status_history__changed_by'),
        pk=reservation_id
    )

    old_status = reservation.status  # pour détecter le changement

    if request.method == "POST":
        form = ReservationAdminForm(request.POST, instance=reservation)
        if form.is_valid():
            updated_reservation = form.save()

            new_status = updated_reservation.status
            note = form.cleaned_data.get('admin_comment', '') or ''

            # Historique de statut
            log_reservation_status_change(
                reservation=updated_reservation,
                old_status=old_status,
                new_status=new_status,
                changed_by=request.user,
                note=note
            )

            # Notification au client si statut a changé
            if old_status != new_status:
                notify_user_reservation_status_change(
                    reservation=updated_reservation,
                    old_status=ReservationStatus(old_status).label if old_status else old_status,
                    new_status=ReservationStatus(new_status).label if new_status else new_status,
                    actor=request.user,
                )

            messages.success(request, "Réservation mise à jour avec succès.")
            return redirect('dashboard:reservations_detail', reservation_id=updated_reservation.id)
    else:
        form = ReservationAdminForm(instance=reservation)

    status_history = reservation.status_history.select_related('changed_by').all()

    # --- Extraction des infos client depuis admin_comment (studio) ---
    event_type_label = None
    guests_count = None
    user_message = None

    if reservation.admin_comment:
        lines = reservation.admin_comment.splitlines()

        # Type d'événement & nombre de personnes (lignes simples)
        for line in lines:
            l = line.strip()
            if l.startswith("Type d’événement :") or l.startswith("Type d'événement :"):
                event_type_label = l.split(":", 1)[1].strip()
            elif l.startswith("Nombre de personnes :"):
                guests_count = l.split(":", 1)[1].strip()

        # Message utilisateur : lignes après "Message utilisateur :"
        start_idx = None
        for i, line in enumerate(lines):
            if line.strip().startswith("Message utilisateur"):
                start_idx = i + 1
                break
        if start_idx is not None and start_idx < len(lines):
            user_message = "\n".join(lines[start_idx:]).strip()
            if not user_message:
                user_message = None

    context = {
        "reservation": reservation,
        "form": form,
        "status_history": status_history,
        "event_type_label": event_type_label,
        "guests_count": guests_count,
        "user_message": user_message,
        "unread_notifications_count": Notification.objects.filter(user=request.user, is_read=False).count(),
        "unread_messages_count": Notification.objects.filter(
            user=request.user,
            notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
            is_read=False
        ).count(),
    }
    return render(request, "admin/reservations/detail.html", context)


# ---------- GESTION DES SERVICES ----------

@staff_member_required
def service_list_view(request):
    qs = (
        Service.objects
        .select_related('category')
        .order_by('name')
    )

    # --- Recherche texte ---
    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(
            Q(name__icontains=q) |
            Q(short_description__icontains=q) |
            Q(description__icontains=q) |
            Q(category__name__icontains=q)
        )

    # --- Filtres ---
    selected_category = request.GET.get('category', '')
    selected_service_type = request.GET.get('service_type', '')
    selected_active = request.GET.get('active', '')  # yes / no

    if selected_category:
        qs = qs.filter(category_id=selected_category)

    if selected_service_type:
        qs = qs.filter(service_type=selected_service_type)

    if selected_active == 'yes':
        qs = qs.filter(is_active=True)
    elif selected_active == 'no':
        qs = qs.filter(is_active=False)

    # --- Statistiques simples sur la liste filtrée ---
    total_count = qs.count()
    active_count = qs.filter(is_active=True).count()
    inactive_count = qs.filter(is_active=False).count()

    # --- Pagination ---
    paginator = Paginator(qs, 10)  # 10 services par page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Conserver les filtres dans la pagination & export
    filters_qd = request.GET.copy()
    if 'page' in filters_qd:
        filters_qd.pop('page')
    base_querystring = filters_qd.urlencode()

    categories = ServiceCategory.objects.all().order_by('name')
    service_type_choices = ServiceTypeChoices.choices

    context = {
        "page_obj": page_obj,
        "q": q,
        "selected_category": selected_category,
        "selected_service_type": selected_service_type,
        "selected_active": selected_active,
        "categories": categories,
        "service_type_choices": service_type_choices,
        "base_querystring": base_querystring,

        "total_count": total_count,
        "active_count": active_count,
        "inactive_count": inactive_count,

        "unread_notifications_count": Notification.objects.filter(user=request.user, is_read=False).count(),
        "unread_messages_count": Notification.objects.filter(
            user=request.user,
            notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
            is_read=False
        ).count(),
    }
    return render(request, "admin/services/list.html", context)


@staff_member_required
def service_detail_view(request, service_id):
    service = get_object_or_404(
        Service.objects.select_related('category'),
        pk=service_id
    )

    context = {
        "service": service,
        "unread_notifications_count": Notification.objects.filter(user=request.user, is_read=False).count(),
        "unread_messages_count": Notification.objects.filter(
            user=request.user,
            notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
            is_read=False
        ).count(),
    }
    return render(request, "admin/services/detail.html", context)


@staff_member_required
def service_export_excel_view(request):
    """
    Exporte la liste des services filtrée au format Excel (XLSX).
    """
    qs = Service.objects.select_related('category').order_by('name')

    # mêmes filtres/recherche que la liste
    q = request.GET.get('q', '').strip()
    selected_category = request.GET.get('category', '')
    selected_service_type = request.GET.get('service_type', '')
    selected_active = request.GET.get('active', '')

    if q:
        qs = qs.filter(
            Q(name__icontains=q) |
            Q(short_description__icontains=q) |
            Q(description__icontains=q) |
            Q(category__name__icontains=q)
        )

    if selected_category:
        qs = qs.filter(category_id=selected_category)

    if selected_service_type:
        qs = qs.filter(service_type=selected_service_type)

    if selected_active == 'yes':
        qs = qs.filter(is_active=True)
    elif selected_active == 'no':
        qs = qs.filter(is_active=False)

    # Création du classeur
    wb = Workbook()
    ws = wb.active
    ws.title = "Services"

    headers = [
        "ID",
        "Nom",
        "Catégorie",
        "Type",
        "Prix de base",
        "Durée min (min)",
        "Durée max (min)",
        "Lieu de prestation",
        "Complexité",
        "Nécessite studio",
        "Nécessite location matériel",
        "Actif",
        "Créé le",
    ]
    ws.append(headers)

    for s in qs:
        ws.append([
            s.id,
            s.name,
            s.category.name if s.category else "",
            s.get_service_type_display(),
            float(s.base_price),
            s.duration_min_minutes or "",
            s.duration_max_minutes or "",
            s.get_location_type_display(),
            s.get_difficulty_level_display(),
            "Oui" if s.requires_studio else "Non",
            "Oui" if s.requires_equipment_rental else "Non",
            "Oui" if s.is_active else "Non",
            s.created_at.strftime("%d/%m/%Y %H:%M") if s.created_at else "",
        ])

    # Style des en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="111827")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Largeur des colonnes
    column_widths = [6, 26, 20, 18, 14, 12, 12, 18, 16, 16, 20, 10, 18]
    for i, column_cells in enumerate(ws.columns, start=1):
        try:
            ws.column_dimensions[column_cells[0].column_letter].width = column_widths[i-1]
        except IndexError:
            pass

    # Réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    timestamp = timezone.now().strftime("%Y%m%d_%H%M")
    filename = f"services_oloustream_{timestamp}.xlsx"
    response['Content-Disposition'] = f'attachment; filename=\"{filename}\"'

    wb.save(response)
    return response


@staff_member_required
def service_create_view(request):
    if request.method == "POST":
        form = ServiceForm(request.POST, request.FILES, instance=service)  # <-- + request.FILES
        if form.is_valid():
            service = form.save()
            messages.success(request, f"Service {service.name} créé avec succès.")
            return redirect('dashboard:services_list')
    else:
        form = ServiceForm()

    context = {
        "form": form,
        "service": None,
        "unread_notifications_count": Notification.objects.filter(user=request.user, is_read=False).count(),
        "unread_messages_count": Notification.objects.filter(
            user=request.user,
            notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
            is_read=False
        ).count(),
    }
    return render(request, "admin/services/form.html", context)


@staff_member_required
def service_update_view(request, service_id):
    service = get_object_or_404(Service, pk=service_id)

    if request.method == "POST":
        form = ServiceForm(request.POST, request.FILES, instance=service)
        if form.is_valid():
            form.save()
            messages.success(request, "Service mis à jour avec succès.")
            return redirect('dashboard:services_list')
    else:
        form = ServiceForm(instance=service)

    context = {
        "form": form,
        "service": service,
        "unread_notifications_count": Notification.objects.filter(user=request.user, is_read=False).count(),
        "unread_messages_count": Notification.objects.filter(
            user=request.user,
            notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
            is_read=False
        ).count(),
    }
    return render(request, "admin/services/form.html", context)


@staff_member_required
def service_delete_view(request, service_id):
    service = get_object_or_404(Service, pk=service_id)

    if request.method == "POST":
        name = service.name
        service.delete()
        messages.success(request, f"Service {name} supprimé.")
        return redirect('dashboard:services_list')

    context = {
        "service": service,
        "unread_notifications_count": Notification.objects.filter(user=request.user, is_read=False).count(),
        "unread_messages_count": Notification.objects.filter(
            user=request.user,
            notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
            is_read=False
        ).count(),
    }
    return render(request, "admin/services/confirm_delete.html", context)


# ---------- GESTION DES OFFRES ----------

@staff_member_required
def offer_list_view(request):
    qs = (
        Offer.objects
        .select_related('service')
        .order_by('-start_date')
    )

    today = timezone.now().date()

    # --- Recherche texte ---
    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(
            Q(title__icontains=q) |
            Q(description__icontains=q) |
            Q(service__name__icontains=q)
        )

    # --- Filtres ---
    selected_service = request.GET.get('service', '')
    selected_active = request.GET.get('active', '')  # yes / no
    selected_period = request.GET.get('period', '')  # current / upcoming / expired

    if selected_service:
        qs = qs.filter(service_id=selected_service)

    if selected_active == 'yes':
        qs = qs.filter(is_active=True)
    elif selected_active == 'no':
        qs = qs.filter(is_active=False)

    if selected_period == 'current':
        qs = qs.filter(start_date__lte=today, end_date__gte=today)
    elif selected_period == 'upcoming':
        qs = qs.filter(start_date__gt=today)
    elif selected_period == 'expired':
        qs = qs.filter(end_date__lt=today)

    # --- Statistiques sur la liste filtrée ---
    total_count = qs.count()
    active_count = qs.filter(is_active=True).count()
    inactive_count = qs.filter(is_active=False).count()
    current_count = qs.filter(is_active=True, start_date__lte=today, end_date__gte=today).count()
    upcoming_count = qs.filter(start_date__gt=today).count()
    expired_count = qs.filter(end_date__lt=today).count()

    # --- Pagination ---
    paginator = Paginator(qs, 10)  # 10 offres par page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Conserver les filtres pour pagination & export
    filters_qd = request.GET.copy()
    if 'page' in filters_qd:
        filters_qd.pop('page')
    base_querystring = filters_qd.urlencode()

    services = Service.objects.filter(is_active=True).order_by('name')

    context = {
        "page_obj": page_obj,
        "q": q,
        "selected_service": selected_service,
        "selected_active": selected_active,
        "selected_period": selected_period,
        "services": services,
        "base_querystring": base_querystring,

        "total_count": total_count,
        "active_count": active_count,
        "inactive_count": inactive_count,
        "current_count": current_count,
        "upcoming_count": upcoming_count,
        "expired_count": expired_count,

        "unread_notifications_count": Notification.objects.filter(user=request.user, is_read=False).count(),
        "unread_messages_count": Notification.objects.filter(
            user=request.user,
            notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
            is_read=False
        ).count(),
    }
    return render(request, "admin/offers/list.html", context)


@staff_member_required
def offer_export_excel_view(request):
    """
    Exporte la liste des offres filtrée au format Excel (XLSX).
    """
    qs = Offer.objects.select_related('service').order_by('-start_date')
    today = timezone.now().date()

    # mêmes filtres que la liste
    q = request.GET.get('q', '').strip()
    selected_service = request.GET.get('service', '')
    selected_active = request.GET.get('active', '')
    selected_period = request.GET.get('period', '')

    if q:
        qs = qs.filter(
            Q(title__icontains=q) |
            Q(description__icontains=q) |
            Q(service__name__icontains=q)
        )

    if selected_service:
        qs = qs.filter(service_id=selected_service)

    if selected_active == 'yes':
        qs = qs.filter(is_active=True)
    elif selected_active == 'no':
        qs = qs.filter(is_active=False)

    if selected_period == 'current':
        qs = qs.filter(start_date__lte=today, end_date__gte=today)
    elif selected_period == 'upcoming':
        qs = qs.filter(start_date__gt=today)
    elif selected_period == 'expired':
        qs = qs.filter(end_date__lt=today)

    wb = Workbook()
    ws = wb.active
    ws.title = "Offres"

    headers = [
        "ID",
        "Titre",
        "Service",
        "Réduction (%)",
        "Date début",
        "Date fin",
        "Période",
        "Active",
        "Description (200 chars)",
    ]
    ws.append(headers)

    for o in qs:
        if o.start_date and o.end_date:
            if o.start_date <= today <= o.end_date:
                period_label = "En cours"
            elif o.start_date > today:
                period_label = "À venir"
            else:
                period_label = "Expirée"
        else:
            period_label = ""

        ws.append([
            o.id,
            o.title,
            o.service.name if o.service else "",
            o.discount_percent,
            o.start_date.strftime("%d/%m/%Y") if o.start_date else "",
            o.end_date.strftime("%d/%m/%Y") if o.end_date else "",
            period_label,
            "Oui" if o.is_active else "Non",
            (o.description or "")[:200],
        ])

    # Style en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="111827")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Largeur colonnes
    column_widths = [6, 26, 24, 14, 14, 14, 14, 10, 40]
    for i, column_cells in enumerate(ws.columns, start=1):
        try:
            ws.column_dimensions[column_cells[0].column_letter].width = column_widths[i-1]
        except IndexError:
            pass

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    timestamp = timezone.now().strftime("%Y%m%d_%H%M")
    filename = f"offres_oloustream_{timestamp}.xlsx"
    response['Content-Disposition'] = f'attachment; filename=\"{filename}\"'

    wb.save(response)
    return response


@staff_member_required
def offer_detail_view(request, offer_id):
    offer = get_object_or_404(
        Offer.objects.select_related('service'),
        pk=offer_id
    )

    today = timezone.now().date()
    if offer.start_date and offer.end_date:
        if offer.start_date <= today <= offer.end_date:
            period_label = "En cours"
        elif offer.start_date > today:
            period_label = "À venir"
        else:
            period_label = "Expirée"
    else:
        period_label = "Non définie"

    context = {
        "offer": offer,
        "period_label": period_label,
        "unread_notifications_count": Notification.objects.filter(user=request.user, is_read=False).count(),
        "unread_messages_count": Notification.objects.filter(
            user=request.user,
            notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
            is_read=False
        ).count(),
    }
    return render(request, "admin/offers/detail.html", context)


@staff_member_required
def offer_create_view(request):
    if request.method == "POST":
        form = OfferForm(request.POST)
        if form.is_valid():
            offer = form.save()
            messages.success(request, f"Offre {offer.title} créée avec succès.")
            return redirect('dashboard:offers_list')
    else:
        form = OfferForm()

    context = {
        "form": form,
        "offer": None,
        "unread_notifications_count": Notification.objects.filter(user=request.user, is_read=False).count(),
        "unread_messages_count": Notification.objects.filter(
            user=request.user,
            notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
            is_read=False
        ).count(),
    }
    return render(request, "admin/offers/form.html", context)


@staff_member_required
def offer_update_view(request, offer_id):
    offer = get_object_or_404(Offer, pk=offer_id)

    if request.method == "POST":
        form = OfferForm(request.POST, instance=offer)
        if form.is_valid():
            form.save()
            messages.success(request, "Offre mise à jour avec succès.")
            return redirect('dashboard:offers_list')
    else:
        form = OfferForm(instance=offer)

    context = {
        "form": form,
        "offer": offer,
        "unread_notifications_count": Notification.objects.filter(user=request.user, is_read=False).count(),
        "unread_messages_count": Notification.objects.filter(
            user=request.user,
            notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
            is_read=False
        ).count(),
    }
    return render(request, "admin/offers/form.html", context)


@staff_member_required
def offer_delete_view(request, offer_id):
    offer = get_object_or_404(Offer, pk=offer_id)

    if request.method == "POST":
        title = offer.title
        offer.delete()
        messages.success(request, f"Offre {title} supprimée.")
        return redirect('dashboard:offers_list')

    context = {
        "offer": offer,
        "unread_notifications_count": Notification.objects.filter(user=request.user, is_read=False).count(),
        "unread_messages_count": Notification.objects.filter(
            user=request.user,
            notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
            is_read=False
        ).count(),
    }
    return render(request, "admin/offers/confirm_delete.html", context)


# ---------- GESTION DES FORMATIONS ----------

@staff_member_required
def training_list_view(request):
    qs = Training.objects.select_related('category').order_by('title')

    # Recherche
    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(
            Q(title__icontains=q) |
            Q(short_description__icontains=q) |
            Q(description__icontains=q) |
            Q(category__name__icontains=q)
        )

    # Filtres
    selected_category = request.GET.get('category', '')
    selected_level = request.GET.get('level', '')
    selected_mode = request.GET.get('mode', '')
    selected_active = request.GET.get('active', '')

    if selected_category:
        qs = qs.filter(category_id=selected_category)

    if selected_level:
        qs = qs.filter(level=selected_level)

    if selected_mode:
        qs = qs.filter(mode=selected_mode)

    if selected_active == 'yes':
        qs = qs.filter(is_active=True)
    elif selected_active == 'no':
        qs = qs.filter(is_active=False)

    # Stats
    total_count = qs.count()
    active_count = qs.filter(is_active=True).count()
    inactive_count = qs.filter(is_active=False).count()

    # Pagination
    paginator = Paginator(qs, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    filters_qd = request.GET.copy()
    if 'page' in filters_qd:
        filters_qd.pop('page')
    base_querystring = filters_qd.urlencode()

    categories = TrainingCategory.objects.all().order_by('name')
    levels = TrainingLevelChoices.choices
    modes = TrainingModeChoices.choices

    context = {
        "page_obj": page_obj,
        "q": q,
        "selected_category": selected_category,
        "selected_level": selected_level,
        "selected_mode": selected_mode,
        "selected_active": selected_active,
        "categories": categories,
        "levels": levels,
        "modes": modes,
        "base_querystring": base_querystring,
        "total_count": total_count,
        "active_count": active_count,
        "inactive_count": inactive_count,
        "unread_notifications_count": Notification.objects.filter(user=request.user, is_read=False).count(),
        "unread_messages_count": Notification.objects.filter(
            user=request.user,
            notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
            is_read=False
        ).count(),
    }
    return render(request, "admin/trainings/list.html", context)

@staff_member_required
def training_export_excel_view(request):
    qs = Training.objects.select_related('category').order_by('title')

    # mêmes filtres que la liste
    q = request.GET.get('q', '').strip()
    selected_category = request.GET.get('category', '')
    selected_level = request.GET.get('level', '')
    selected_mode = request.GET.get('mode', '')
    selected_active = request.GET.get('active', '')

    if q:
        qs = qs.filter(
            Q(title__icontains=q) |
            Q(short_description__icontains=q) |
            Q(description__icontains=q) |
            Q(category__name__icontains=q)
        )
    if selected_category:
        qs = qs.filter(category_id=selected_category)
    if selected_level:
        qs = qs.filter(level=selected_level)
    if selected_mode:
        qs = qs.filter(mode=selected_mode)
    if selected_active == 'yes':
        qs = qs.filter(is_active=True)
    elif selected_active == 'no':
        qs = qs.filter(is_active=False)

    wb = Workbook()
    ws = wb.active
    ws.title = "Formations"

    headers = [
        "ID",
        "Titre",
        "Catégorie",
        "Niveau",
        "Mode",
        "Lieu",
        "Durée (h)",
        "Prix (F CFA)",
        "Certification",
        "Actif",
        "Date début",
        "Date fin",
    ]
    ws.append(headers)

    for t in qs:
        ws.append([
            t.id,
            t.title,
            t.category.name if t.category else "",
            t.get_level_display(),
            t.get_mode_display(),
            t.location,
            t.duration_hours or "",
            float(t.price) if t.price is not None else None,
            "Oui" if t.certification else "Non",
            "Oui" if t.is_active else "Non",
            t.start_date.strftime("%d/%m/%Y") if t.start_date else "",
            t.end_date.strftime("%d/%m/%Y") if t.end_date else "",
        ])

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="111827")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    col_widths = [6, 26, 20, 16, 16, 20, 10, 14, 12, 10, 14, 14]
    for i, col in enumerate(ws.columns, start=1):
        try:
            ws.column_dimensions[col[0].column_letter].width = col_widths[i-1]
        except IndexError:
            pass

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    timestamp = timezone.now().strftime("%Y%m%d_%H%M")
    filename = f"formations_oloustream_{timestamp}.xlsx"
    response['Content-Disposition'] = f'attachment; filename=\"{filename}\"'

    wb.save(response)
    return response


@staff_member_required
def training_create_view(request):
    if request.method == "POST":
        form = TrainingForm(request.POST, request.FILES)
        if form.is_valid():
            training = form.save()
            messages.success(request, f"Formation {training.title} créée avec succès.")
            return redirect('dashboard:trainings_list')
    else:
        form = TrainingForm()

    context = {
        "form": form,
        "training": None,
        "unread_notifications_count": Notification.objects.filter(user=request.user, is_read=False).count(),
        "unread_messages_count": Notification.objects.filter(
            user=request.user,
            notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
            is_read=False
        ).count(),
    }
    return render(request, "admin/trainings/form.html", context)

@staff_member_required
def training_update_view(request, training_id):
    training = get_object_or_404(Training, pk=training_id)

    if request.method == "POST":
        form = TrainingForm(request.POST, request.FILES, instance=training)
        if form.is_valid():
            form.save()
            messages.success(request, "Formation mise à jour avec succès.")
            return redirect('dashboard:trainings_list')
    else:
        form = TrainingForm(instance=training)

    context = {
        "form": form,
        "training": training,
       "unread_notifications_count": Notification.objects.filter(user=request.user, is_read=False).count(),
        "unread_messages_count": Notification.objects.filter(
            user=request.user,
            notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
            is_read=False
        ).count(),
    }
    return render(request, "admin/trainings/form.html", context)



@staff_member_required
def training_delete_view(request, training_id):
    training = get_object_or_404(Training, pk=training_id)

    if request.method == "POST":
        title = training.title
        training.delete()
        messages.success(request, f"Formation {title} supprimée.")
        return redirect('dashboard:trainings_list')

    context = {
        "training": training,
        "unread_notifications_count": Notification.objects.filter(user=request.user, is_read=False).count(),
        "unread_messages_count": Notification.objects.filter(
            user=request.user,
            notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
            is_read=False
        ).count(),
    }
    return render(request, "admin/trainings/confirm_delete.html", context)

@staff_member_required
def training_detail_view(request, training_id):
    training = get_object_or_404(
        Training.objects.select_related('category'),
        pk=training_id
    )

    today = timezone.now().date()
    if training.start_date and training.end_date:
        if training.start_date <= today <= training.end_date:
            period_label = "En cours"
        elif training.start_date > today:
            period_label = "À venir"
        else:
            period_label = "Terminée"
    else:
        period_label = "Non planifiée"

    context = {
        "training": training,
        "period_label": period_label,
        "unread_notifications_count": Notification.objects.filter(user=request.user, is_read=False).count(),
        "unread_messages_count": Notification.objects.filter(
            user=request.user,
            notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
            is_read=False
        ).count(),
    }
    return render(request, "admin/trainings/detail.html", context)


# ---------- GESTION DES PARTENAIRES ----------

@staff_member_required
def partner_list_view(request):
    qs = Partner.objects.all().order_by('name')

    # Recherche
    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(
            Q(name__icontains=q) |
            Q(website__icontains=q)
        )

    # Filtre actif
    selected_active = request.GET.get('active', '')  # yes / no
    if selected_active == 'yes':
        qs = qs.filter(active=True)
    elif selected_active == 'no':
        qs = qs.filter(active=False)

    # Stats
    total_count = qs.count()
    active_count = qs.filter(active=True).count()
    inactive_count = qs.filter(active=False).count()

    # Pagination
    paginator = Paginator(qs, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Querystring sans page
    filters_qd = request.GET.copy()
    if 'page' in filters_qd:
        filters_qd.pop('page')
    base_querystring = filters_qd.urlencode()

    context = {
        "page_obj": page_obj,
        "q": q,
        "selected_active": selected_active,
        "base_querystring": base_querystring,
        "total_count": total_count,
        "active_count": active_count,
        "inactive_count": inactive_count,
        "unread_notifications_count": Notification.objects.filter(user=request.user, is_read=False).count(),
        "unread_messages_count": Notification.objects.filter(
            user=request.user,
            notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
            is_read=False
        ).count(),
    }
    return render(request, "admin/partners/list.html", context)

@staff_member_required
def partner_export_excel_view(request):
    """
    Exporte la liste des partenaires filtrée au format Excel (XLSX).
    """
    qs = Partner.objects.all().order_by('name')

    q = request.GET.get('q', '').strip()
    selected_active = request.GET.get('active', '')

    if q:
        qs = qs.filter(
            Q(name__icontains=q) |
            Q(website__icontains=q)
        )

    if selected_active == 'yes':
        qs = qs.filter(active=True)
    elif selected_active == 'no':
        qs = qs.filter(active=False)

    wb = Workbook()
    ws = wb.active
    ws.title = "Partenaires"

    headers = [
        "ID",
        "Nom",
        "Site web",
        "Actif",
    ]
    ws.append(headers)

    for p in qs:
        ws.append([
            p.id,
            p.name,
            p.website,
            "Oui" if p.active else "Non",
        ])

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="111827")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    col_widths = [6, 26, 30, 10]
    for i, col in enumerate(ws.columns, start=1):
        try:
            ws.column_dimensions[col[0].column_letter].width = col_widths[i-1]
        except IndexError:
            pass

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    timestamp = timezone.now().strftime("%Y%m%d_%H%M")
    filename = f"partenaires_oloustream_{timestamp}.xlsx"
    response['Content-Disposition'] = f'attachment; filename=\"{filename}\"'

    wb.save(response)
    return response

@staff_member_required
def partner_detail_view(request, partner_id):
    partner = get_object_or_404(Partner, pk=partner_id)

    context = {
        "partner": partner,
        "unread_notifications_count": Notification.objects.filter(user=request.user, is_read=False).count(),
        "unread_messages_count": Notification.objects.filter(
            user=request.user,
            notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
            is_read=False
        ).count(),
    }
    return render(request, "admin/partners/detail.html", context)


@staff_member_required
def partner_create_view(request):
    if request.method == "POST":
        form = PartnerForm(request.POST, request.FILES)
        if form.is_valid():
            partner = form.save()
            messages.success(request, f"Partenaire {partner.name} créé avec succès.")
            return redirect('dashboard:partners_list')
    else:
        form = PartnerForm()

    context = {
        "form": form,
        "partner": None,
        "unread_notifications_count": Notification.objects.filter(user=request.user, is_read=False).count(),
        "unread_messages_count": Notification.objects.filter(
            user=request.user,
            notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
            is_read=False
        ).count(),
    }
    return render(request, "admin/partners/form.html", context)



@staff_member_required
def partner_update_view(request, partner_id):
    partner = get_object_or_404(Partner, pk=partner_id)

    if request.method == "POST":
        form = PartnerForm(request.POST, request.FILES, instance=partner)
        if form.is_valid():
            form.save()
            messages.success(request, "Partenaire mis à jour avec succès.")
            return redirect('dashboard:partners_list')
    else:
        form = PartnerForm(instance=partner)

    context = {
        "form": form,
        "partner": partner,
        "unread_notifications_count": Notification.objects.filter(user=request.user, is_read=False).count(),
        "unread_messages_count": Notification.objects.filter(
            user=request.user,
            notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
            is_read=False
        ).count(),
    }
    return render(request, "admin/partners/form.html", context)


@staff_member_required
def partner_delete_view(request, partner_id):
    partner = get_object_or_404(Partner, pk=partner_id)

    if request.method == "POST":
        name = partner.name
        partner.delete()
        messages.success(request, f"Partenaire {name} supprimé.")
        return redirect('dashboard:partners_list')

    context = {
        "partner": partner,
        "unread_notifications_count": Notification.objects.filter(user=request.user, is_read=False).count(),
        "unread_messages_count": Notification.objects.filter(
            user=request.user,
            notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
            is_read=False
        ).count(),
    }
    return render(request, "admin/partners/confirm_delete.html", context)


# ---------- GESTION DES STUDIOS (DASHBOARD) ----------

@staff_member_required
def studio_list_view(request):
    qs = Studio.objects.all().order_by("name")

    context = {
        "studios": qs,
        "unread_notifications_count": Notification.objects.filter(
            user=request.user, is_read=False
        ).count(),
        "unread_messages_count": Notification.objects.filter(
            user=request.user,
            notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
            is_read=False,
        ).count(),
    }
    return render(request, "admin/studios/list.html", context)


@staff_member_required
def studio_create_view(request):
    if request.method == "POST":
        form = StudioForm(request.POST, request.FILES)
        if form.is_valid():
            studio = form.save()
            messages.success(
                request,
                f"Le studio « {studio.name} » a été créé avec succès."
            )
            return redirect("dashboard:studios_list")
    else:
        form = StudioForm()

    context = {
        "form": form,
        "studio": None,
        "unread_notifications_count": Notification.objects.filter(
            user=request.user, is_read=False
        ).count(),
        "unread_messages_count": Notification.objects.filter(
            user=request.user,
            notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
            is_read=False,
        ).count(),
    }
    return render(request, "admin/studios/form.html", context)


@staff_member_required
def studio_update_view(request, studio_id):
    studio = get_object_or_404(Studio, pk=studio_id)

    if request.method == "POST":
        form = StudioForm(request.POST, request.FILES, instance=studio)
        if form.is_valid():
            studio = form.save()
            messages.success(
                request,
                f"Le studio « {studio.name} » a été mis à jour."
            )
            return redirect("dashboard:studios_detail", studio_id=studio.pk)
    else:
        form = StudioForm(instance=studio)

    context = {
        "form": form,
        "studio": studio,
        "unread_notifications_count": Notification.objects.filter(
            user=request.user, is_read=False
        ).count(),
        "unread_messages_count": Notification.objects.filter(
            user=request.user,
            notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
            is_read=False,
        ).count(),
    }
    return render(request, "admin/studios/form.html", context)


@staff_member_required
def studio_detail_view(request, studio_id):
    studio = get_object_or_404(Studio, pk=studio_id)

    context = {
        "studio": studio,
        "unread_notifications_count": Notification.objects.filter(
            user=request.user, is_read=False
        ).count(),
        "unread_messages_count": Notification.objects.filter(
            user=request.user,
            notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
            is_read=False,
        ).count(),
    }
    return render(request, "admin/studios/detail.html", context)


@staff_member_required
def studio_delete_view(request, studio_id):
    studio = get_object_or_404(Studio, pk=studio_id)

    if request.method == "POST":
        name = studio.name
        studio.delete()
        messages.success(request, f"Le studio « {name} » a été supprimé.")
        return redirect("dashboard:studios_list")

    context = {
        "studio": studio,
        "unread_notifications_count": Notification.objects.filter(
            user=request.user, is_read=False
        ).count(),
        "unread_messages_count": Notification.objects.filter(
            user=request.user,
            notification_type=NotificationTypeChoices.MESSAGE_RECEIVED,
            is_read=False,
        ).count(),
    }
    return render(request, "admin/studios/confirm_delete.html", context)



#=====================================================================================================================================================================================
# apps/dashboard/views.py (à ajouter)


def _status_label(value: str) -> str:
    try:
        return ReservationStatus(value).label
    except Exception:
        return value


@staff_member_required
@require_POST
def reservation_set_status_view(request, reservation_id):
    """
    Action rapide POST :
    - CONFIRMED / REJECTED / COMPLETED / CANCELLED
    """
    reservation = Reservation.objects.select_related("user").get(pk=reservation_id)

    new_status = request.POST.get("status", "").strip()
    next_url = request.POST.get("next") or request.META.get("HTTP_REFERER") or "/dashboard/reservations/"

    allowed = {
        ReservationStatus.CONFIRMED,
        ReservationStatus.REJECTED,
        ReservationStatus.COMPLETED,
        ReservationStatus.CANCELLED,
    }

    if new_status not in [s.value for s in allowed]:
        messages.error(request, "Statut invalide.")
        return redirect(next_url)

    old_status = reservation.status

    if old_status == new_status:
        messages.info(request, "Aucun changement : la réservation a déjà ce statut.")
        return redirect(next_url)

    # ✅ Optionnel (recommandé) : empêcher de confirmer une réservation déjà passée
    if new_status == ReservationStatus.CONFIRMED and reservation.end_datetime and reservation.end_datetime < timezone.now():
        messages.error(request, "Impossible de confirmer : ce créneau est déjà passé.")
        return redirect(next_url)

    reservation.status = new_status
    reservation.save(update_fields=["status"])

    # Historique
    log_reservation_status_change(
        reservation=reservation,
        old_status=old_status,
        new_status=new_status,
        changed_by=request.user,
        note=f"Changement rapide de statut -> {new_status}",
    )

    # Notification client
    notify_user_reservation_status_change(
        reservation=reservation,
        old_status=_status_label(old_status),
        new_status=_status_label(new_status),
        actor=request.user,
    )

    send_reservation_status_changed_email(
    request=request,
    reservation=reservation,
    old_status_label=_status_label(old_status),
    new_status_label=_status_label(new_status),
    admin_note = reservation.admin_comment or ""
    )

    messages.success(request, f"Statut mis à jour : {_status_label(new_status)}")
    return redirect(next_url)






#============================================================= pour la partie business partners ====================================
# ============================================================
# VUES PARTENAIRES D'AFFAIRES
# ============================================================

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Sum, Count, Q
from django.utils import timezone
from django.http import JsonResponse

from apps.business_partners.models import (
    PartnerApplication, BusinessPartner, Contract, 
    CommissionPayment, Region
)
from apps.business_partners.services import activate_partner


def is_staff(user):
    """Vérifie que l'utilisateur est staff"""
    return user.is_staff


# ==================== CANDIDATURES ====================

@login_required
@user_passes_test(is_staff)
def partner_applications_list(request):
    """Liste des candidatures partenaires"""
    applications = PartnerApplication.objects.all().order_by('-created_at')
    
    # Filtres
    status = request.GET.get('status')
    city = request.GET.get('city')
    network = request.GET.get('network')
    search = request.GET.get('search')
    
    if status:
        applications = applications.filter(status=status)
    if city:
        applications = applications.filter(city_id=city)
    if network:
        applications = applications.filter(network_strength=network)
    if search:
        applications = applications.filter(
            Q(full_name__icontains=search) |
            Q(phone__icontains=search) |
            Q(email__icontains=search)
        )
    
    # Statistiques
    stats = {
        'total': PartnerApplication.objects.count(),
        'pending': PartnerApplication.objects.filter(status='pending').count(),
        'reviewing': PartnerApplication.objects.filter(status='reviewing').count(),
        'approved': PartnerApplication.objects.filter(status='approved').count(),
        'rejected': PartnerApplication.objects.filter(status='rejected').count(),
    }
    
    # Pagination
    paginator = Paginator(applications, 15)
    page = request.GET.get('page')
    applications = paginator.get_page(page)
    
    context = {
        'applications': applications,
        'stats': stats,
        'regions': Region.objects.all(),
        'status_choices': PartnerApplication.STATUS_CHOICES,
        'network_choices': PartnerApplication.NETWORK_STRENGTH,
        'current_status': status,
        'current_city': city,
        'current_network': network,
        'search_query': search,
    }
    return render(request, 'admin/business_partners/applications_list.html', context)


@login_required
@user_passes_test(is_staff)
def partner_application_detail(request, pk):
    """Détail d'une candidature"""
    application = get_object_or_404(PartnerApplication, pk=pk)
    
    context = {
        'application': application,
    }
    return render(request, 'admin/business_partners/application_detail.html', context)


@login_required
@user_passes_test(is_staff)
def partner_application_approve(request, pk):
    """Approuver une candidature"""
    application = get_object_or_404(PartnerApplication, pk=pk)
    
    if application.status == 'approved':
        messages.warning(request, "⚠️ Cette candidature est déjà approuvée.")
    else:
        try:
            partner = activate_partner(application, request.user)
            messages.success(
                request,
                f"✅ Partenaire activé ! Code : {partner.partner_code}. "
                f"Un email avec les identifiants a été envoyé à {partner.user.email}"
            )
            return redirect('dashboard:business_partner_detail', pk=partner.pk)
        except Exception as e:
            messages.error(request, f"❌ Erreur : {str(e)}")
    
    return redirect('dashboard:partner_applications_list')


@login_required
@user_passes_test(is_staff)
def partner_application_reject(request, pk):
    """Rejeter une candidature"""
    application = get_object_or_404(PartnerApplication, pk=pk)
    
    reason = request.POST.get('reason', '')
    
    application.status = 'rejected'
    application.reviewed_by = request.user
    application.reviewed_at = timezone.now()
    application.internal_notes = f"Rejetée le {timezone.now().strftime('%d/%m/%Y %H:%M')}\nRaison : {reason}"
    application.save()
    
    messages.warning(request, f"❌ Candidature de {application.full_name} rejetée.")
    return redirect('dashboard:partner_applications_list')


@login_required
@user_passes_test(is_staff)
def partner_application_update_status(request, pk):
    """Mettre à jour le statut d'une candidature"""
    application = get_object_or_404(PartnerApplication, pk=pk)
    
    new_status = request.POST.get('status')
    if new_status in dict(PartnerApplication.STATUS_CHOICES):
        application.status = new_status
        application.save()
        messages.success(request, f"✅ Statut mis à jour : {application.get_status_display()}")
    
    return redirect('dashboard:partner_application_detail', pk=pk)


# ==================== PARTENAIRES ACTIFS ====================

@login_required
@user_passes_test(is_staff)
def business_partners_list(request):
    """Liste des partenaires actifs"""
    partners = BusinessPartner.objects.all().select_related(
        'user', 'application', 'application__city'
    ).order_by('-total_revenue')
    
    # Filtres
    active_filter = request.GET.get('active')
    city_filter = request.GET.get('city')
    search = request.GET.get('search')
    
    if active_filter == 'true':
        partners = partners.filter(is_active=True)
    elif active_filter == 'false':
        partners = partners.filter(is_active=False)
    
    if city_filter:
        partners = partners.filter(application__city_id=city_filter)
    
    if search:
        partners = partners.filter(
            Q(partner_code__icontains=search) |
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(user__email__icontains=search)
        )
    
    # Statistiques globales
    stats = BusinessPartner.objects.aggregate(
        total_partners=Count('id'),
        active_partners=Count('id', filter=Q(is_active=True)),
        total_revenue=Sum('total_revenue'),
        total_commission=Sum('total_commission_earned'),
        total_paid=Sum('total_commission_paid'),
    )
    stats['pending_commission'] = (stats['total_commission'] or 0) - (stats['total_paid'] or 0)
    
    # Pagination
    paginator = Paginator(partners, 15)
    page = request.GET.get('page')
    partners = paginator.get_page(page)
    
    context = {
        'partners': partners,
        'stats': stats,
        'regions': Region.objects.all(),
        'current_active': active_filter,
        'current_city': city_filter,
        'search_query': search,
    }
    return render(request, 'admin/business_partners/partners_list.html', context)


@login_required
@user_passes_test(is_staff)
def business_partner_detail(request, pk):
    """Détail d'un partenaire"""
    partner = get_object_or_404(BusinessPartner, pk=pk)
    
    # Contrats du partenaire
    contracts = Contract.objects.filter(partner=partner).order_by('-created_at')[:10]
    
    # Paiements du partenaire
    payments = CommissionPayment.objects.filter(partner=partner).order_by('-paid_at')[:5]
    
    # Statistiques
    contracts_stats = Contract.objects.filter(partner=partner).aggregate(
        total=Count('id'),
        validated=Count('id', filter=Q(status='validated')),
        completed=Count('id', filter=Q(status='completed')),
        pending=Count('id', filter=Q(status='pending')),
    )
    
    context = {
        'partner': partner,
        'contracts': contracts,
        'payments': payments,
        'contracts_stats': contracts_stats,
    }
    return render(request, 'admin/business_partners/partner_detail.html', context)


@login_required
@user_passes_test(is_staff)
def business_partner_toggle_active(request, pk):
    """Activer/Désactiver un partenaire"""
    partner = get_object_or_404(BusinessPartner, pk=pk)
    
    if partner.is_active:
        partner.is_active = False
        partner.suspension_reason = request.POST.get('reason', 'Suspendu par admin')
        messages.warning(request, f"⚠️ Partenaire {partner.partner_code} suspendu.")
    else:
        partner.is_active = True
        partner.suspension_reason = ''
        messages.success(request, f"✅ Partenaire {partner.partner_code} réactivé.")
    
    partner.save()
    return redirect('dashboard:business_partner_detail', pk=pk)


# ==================== CONTRATS ====================

@login_required
@user_passes_test(is_staff)
def partner_contracts_list(request):
    """Liste de tous les contrats"""
    contracts = Contract.objects.all().select_related(
        'partner', 'partner__user'
    ).order_by('-created_at')
    
    # Filtres
    status = request.GET.get('status')
    partner = request.GET.get('partner')
    search = request.GET.get('search')
    
    if status:
        contracts = contracts.filter(status=status)
    if partner:
        contracts = contracts.filter(partner_id=partner)
    if search:
        contracts = contracts.filter(
            Q(client_name__icontains=search) |
            Q(partner__partner_code__icontains=search) |
            Q(service_type__icontains=search)
        )
    
    # Statistiques
    stats = Contract.objects.aggregate(
        total=Count('id'),
        pending=Count('id', filter=Q(status='pending')),
        validated=Count('id', filter=Q(status='validated')),
        completed=Count('id', filter=Q(status='completed')),
        total_amount=Sum('contract_amount'),
        total_commission=Sum('commission_amount'),
    )
    
    # Pagination
    paginator = Paginator(contracts, 20)
    page = request.GET.get('page')
    contracts = paginator.get_page(page)
    
    context = {
        'contracts': contracts,
        'stats': stats,
        'partners': BusinessPartner.objects.filter(is_active=True),
        'status_choices': Contract.STATUS_CHOICES,
        'current_status': status,
        'current_partner': partner,
        'search_query': search,
    }
    return render(request, 'admin/business_partners/contracts_list.html', context)


@login_required
@user_passes_test(is_staff)
def partner_contract_detail(request, pk):
    """Détail d'un contrat"""
    contract = get_object_or_404(Contract, pk=pk)
    
    context = {
        'contract': contract,
    }
    return render(request, 'admin/business_partners/contract_detail.html', context)


@login_required
@user_passes_test(is_staff)
def partner_contract_validate(request, pk):
    """Valider un contrat"""
    contract = get_object_or_404(Contract, pk=pk)
    
    if contract.status != 'pending':
        messages.warning(request, "⚠️ Ce contrat ne peut pas être validé.")
        return redirect('dashboard:partner_contract_detail', pk=pk)
    
    contract.status = 'validated'
    contract.validated_by = request.user
    contract.validated_at = timezone.now()
    contract.save()
    
    # Mettre à jour les stats du partenaire
    partner = contract.partner
    partner.total_contracts += 1
    partner.total_revenue += contract.contract_amount
    partner.total_commission_earned += contract.commission_amount
    partner.last_contract_at = timezone.now()
    partner.save()
    
    messages.success(
        request,
        f"✅ Contrat validé ! Commission de {contract.commission_amount:,.0f} FCFA attribuée à {partner.partner_code}"
    )
    return redirect('dashboard:partner_contracts_list')


@login_required
@user_passes_test(is_staff)
def partner_contract_reject(request, pk):
    """Rejeter/Annuler un contrat"""
    contract = get_object_or_404(Contract, pk=pk)
    
    contract.status = 'cancelled'
    contract.save()
    
    messages.warning(request, f"❌ Contrat de {contract.client_name} annulé.")
    return redirect('dashboard:partner_contracts_list')


# ==================== PAIEMENTS ====================

@login_required
@user_passes_test(is_staff)
def partner_payments_list(request):
    """Liste des paiements de commissions"""
    payments = CommissionPayment.objects.all().select_related(
        'partner', 'partner__user', 'created_by'
    ).order_by('-paid_at')
    
    # Filtres
    partner_filter = request.GET.get('partner')
    method = request.GET.get('method')
    
    if partner_filter:
        payments = payments.filter(partner_id=partner_filter)
    if method:
        payments = payments.filter(payment_method=method)
    
    # Statistiques
    stats = CommissionPayment.objects.aggregate(
        total_payments=Count('id'),
        total_amount=Sum('amount'),
    )
    
    # Pagination
    paginator = Paginator(payments, 20)
    page = request.GET.get('page')
    payments = paginator.get_page(page)
    
    context = {
        'payments': payments,
        'stats': stats,
        'partners': BusinessPartner.objects.filter(is_active=True),
        'method_choices': CommissionPayment.PAYMENT_METHODS,
        'current_partner': partner_filter,
        'current_method': method,
    }
    return render(request, 'admin/business_partners/payments_list.html', context)


@login_required
@user_passes_test(is_staff)
def partner_payment_create(request, partner_pk):
    """Créer un paiement de commission"""
    partner = get_object_or_404(BusinessPartner, pk=partner_pk)
    
    if request.method == 'POST':
        amount = float(request.POST.get('amount', 0))
        method = request.POST.get('payment_method')
        reference = request.POST.get('reference', '')
        notes = request.POST.get('notes', '')
        
        if amount > 0:
            payment = CommissionPayment.objects.create(
                partner=partner,
                amount=amount,
                payment_method=method,
                reference=reference,
                notes=notes,
                created_by=request.user
            )
            
            # Mettre à jour le partenaire
            partner.total_commission_paid += amount
            partner.save()
            
            messages.success(
                request,
                f"✅ Paiement de {amount:,.0f} FCFA enregistré pour {partner.partner_code}"
            )
            return redirect('dashboard:business_partner_detail', pk=partner_pk)
        else:
            messages.error(request, "❌ Le montant doit être supérieur à 0.")
    
    # Contrats non payés
    unpaid_contracts = Contract.objects.filter(
        partner=partner,
        status__in=['validated', 'completed'],
    ).order_by('-created_at')
    
    context = {
        'partner': partner,
        'unpaid_contracts': unpaid_contracts,
        'pending_amount': partner.pending_commission,
        'method_choices': CommissionPayment.PAYMENT_METHODS,
    }
    return render(request, 'admin/business_partners/payment_create.html', context)